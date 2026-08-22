#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_population_age.py — 人口の年齢構成(5歳階級・65歳以上)の整備(issue #28)。

`scripts/fetch_census_age.py` が取得した令和2年国勢調査 人口等基本集計
表2-7(年齢5歳階級)のxlsx(`b02_07`シート)を、339構想区域・47都道府県の
既存人口CSV(`scripts/build_population.py` の出力)に年齢階級列として
突き合わせる。

## 市区町村→構想区域の対応表について

隣リポジトリ(visualize-regional-medical-care-for-2040)の
`iryoken2_A38-20.geojson`(335圏、生のA38属性)は令和2年度時点の二次医療圏
単位で、R7(339区域)とは三重県で粒度が異なる(三重県の旧4圏域がR7で
8区域に細分化されている)。このため:

- 331件(`area_geo_join.csv` で matched)は `A38b_001`(構成市区町村コードの
  カンマ区切り)をそのまま使う。
- 三重県(pref_code=24)の12件(すべて area_geo_join.csv で unmatched)は
  `data/reference/mie_area_municipalities.csv`(三重県公式資料+A38突合で
  作成済みの29市町対応表)を使う。

これにより339区域すべてが漏れなく1つの市区町村集合でカバーされる
(`scripts/build_population.py` は総人口のみでこの年齢内訳を持たない)。

## 地域識別コード(census xlsx col3)の選択

census xlsxの「地域識別コード」列は行の粒度を表す(全国/都道府県/
政令市全体/政令市の区/市区町村(2020年時点)/市区町村(2000年に組み替え))。
このスクリプトは以下を選ぶ:

- `a`: 全国(2020年_地域コード='00000')・都道府県(それ以外)
- `0`, `2`, `3`: 市区町村レベル(政令市の区・通常の市区町村2系統)。
  `1`(政令市全体)は`0`(区)と重複するため使わない。`9`(2000年組み替え)は
  `0`/`2`/`3` と重複する市区町村を含むため使わない。

この選択が正しいことは、`{0,2,3}` の2020年_地域コード集合が
市区町村→構想区域マッピングのmuni_code集合と完全一致すること、および
Σ(この選択の市区町村行) == 全国行(a, 00000) が性別・年齢階級すべてで
一致することで検算する(ハード検算、下記参照)。

## ハード検算(失敗したら非ゼロ終了)

1. 339区域・47都道府県すべてが年齢データを持つ(欠損なし)
2. 性別整合性: 男+女==総数(市区町村・都道府県・全国の各レベル)
3. 全国検算: Σ(選択した市区町村行) == 全国行(性別×年齢階級ごと)、
   かつ Σ(339区域) == 全国行
4. 対応表の網羅性: census の市区町村コード集合 == マッピングのmuni_code集合
   (差分があれば監査CSVに書き、説明できない差分は失敗)
5. pop_total_census と既存 population_2020 の差(相対差0.5%超で失敗)
6. pop_65plus(census '(再掲)65歳以上'列) == Σ(65-69〜100歳以上の5歳階級バンド)。
   census側の集計と5歳階級バンドの合計が食い違えば失敗させる(将来
   e-Statが「(再掲)」の定義を変えた場合に不整合を静かに出荷しないため)。

使い方:
    python scripts/build_population_age.py

終了コード: 成功=0、入力ファイル不在・検算失敗などは非ゼロ。
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import pandas as pd

# 同じ scripts/ に置いた共有モジュールを読む(既存スクリプトと同じ流儀)。
sys.path.insert(0, str(Path(__file__).resolve().parent))
import lib_neighbor_repo  # noqa: E402 (パス追加の後に import する必要がある)

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

DEFAULT_CENSUS_XLSX = Path("data/raw/census_age_2020_table2-7.xlsx")

# 隣リポジトリ側の相対パス。ルートは環境変数 NEIGHBOR_REPO か個別の引数で
# 受け取る(issue #51。開発機の絶対パスを既定値にしない)。
NEIGHBOR_A38_GEOJSON = "data/processed/iryoken2_A38-20.geojson"
NEIGHBOR_AREA_GEO_JOIN = "data/processed/area_geo_join.csv"
NEIGHBOR_MIE_CSV = "data/reference/mie_area_municipalities.csv"

DEFAULT_AREA_CSV = Path("data/processed/population_iryoken2.csv")
DEFAULT_PREF_CSV = Path("data/processed/population_prefecture.csv")
DEFAULT_OUT_DIR = Path("data/processed")

# 既存人口CSVの「基礎列」。冪等性のため、再実行のたびにこの列だけを読み直し
# (年齢列を含んだ状態で読み直すと年齢列が重複する事故を防ぐ)、その後ろに
# 年齢列を必ず同じ列名で付け直す。
BASE_AREA_COLUMNS = ["area_code", "area_name", "pref_code", "pref_name", "population_2020", "source", "retrieved_on"]
BASE_PREF_COLUMNS = ["pref_code", "pref_name", "population_2020", "source", "retrieved_on"]

# census の5歳階級(01〜21)を、この順序で出力列名に対応させる。
# census側のラベル文字列(全角チルダ等)に依存させないよう、番号(01〜21)だけで
# 突き合わせ、列名はここで固定的に定義する(標準的な国勢調査の年齢5歳階級)。
BAND_OUTPUT_COLUMNS = [
    "pop_0_4",
    "pop_5_9",
    "pop_10_14",
    "pop_15_19",
    "pop_20_24",
    "pop_25_29",
    "pop_30_34",
    "pop_35_39",
    "pop_40_44",
    "pop_45_49",
    "pop_50_54",
    "pop_55_59",
    "pop_60_64",
    "pop_65_69",
    "pop_70_74",
    "pop_75_79",
    "pop_80_84",
    "pop_85_89",
    "pop_90_94",
    "pop_95_99",
    "pop_100plus",
]
AGE_EXTRA_COLUMNS = ["pop_65plus", "pop_age_unknown", "pop_total_census"]
AGE_ALL_COLUMNS = BAND_OUTPUT_COLUMNS + AGE_EXTRA_COLUMNS

# census xlsx (b02_07シート) の先頭9列(1〜9列目)の見出し(12行目)。
# 想定どおりの列配置かをここで検証する(構造がズレていたら早期に気づくため)。
EXPECTED_KEY_HEADERS = [
    "国籍総数か日本人",
    "男女",
    "地域識別コード",
    "2000年_都道府県",
    "2000年_地域コード",
    "2000年地域",
    "2020年_都道府県",
    "2020年_地域コード",
    "地域名",
]

SEX_LABEL_MAP = {"0_総数": "total", "1_男": "male", "2_女": "female"}
MUNI_IDENTS = {"0", "2", "3"}

SOURCE_URL = (
    "https://www.e-stat.go.jp/stat-search/file-download?statInfId=000032142410&fileKind=0"
)


def write_csv(path: Path, header: List[str], rows: List[List[object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, lineterminator="\n")
        writer.writerow(header)
        writer.writerows(rows)


class AuditRow:
    __slots__ = ("check", "code", "name", "expected", "actual", "diff", "note")

    def __init__(self, check, code, name, expected, actual, diff, note=""):
        self.check = check
        self.code = code
        self.name = name
        self.expected = expected
        self.actual = actual
        self.diff = diff
        self.note = note

    def as_row(self) -> List[object]:
        return [self.check, self.code, self.name, self.expected, self.actual, self.diff, self.note]


# --- census xlsx 読み込み ----------------------------------------------------


def num(v) -> int:
    """'-'(census の0表記)・空欄を0として扱う。それ以外は int にキャストする。"""
    if v is None or v == "-":
        return 0
    return int(v)


def find_value_columns(ws) -> Dict[str, int]:
    """10行目(見出しコード)・11行目(単位)から「人」単位の値列だけを拾う。

    percentage(％)・平均年齢(歳)の列は同じラベル接頭辞(R1〜R6等)を再利用
    しているため、単位が「人」の列だけに絞ることで人口構成比の列と混同しない。
    """
    label_row = next(ws.iter_rows(min_row=10, max_row=10, values_only=True))
    unit_row = next(ws.iter_rows(min_row=11, max_row=11, values_only=True))
    idx_by_label: Dict[str, int] = {}
    for i, (label, unit) in enumerate(zip(label_row, unit_row)):
        if label and unit == "人":
            idx_by_label[label] = i
    return idx_by_label


def resolve_census_columns(ws) -> Dict[str, object]:
    key_header_row = next(ws.iter_rows(min_row=12, max_row=12, values_only=True))
    actual_key_headers = list(key_header_row[:9])
    if actual_key_headers != EXPECTED_KEY_HEADERS:
        raise ValueError(
            "census xlsxの先頭9列の見出しが想定と異なります(構造が変わった可能性)。\n"
            f"  想定: {EXPECTED_KEY_HEADERS}\n"
            f"  実際: {actual_key_headers}"
        )

    idx_by_label = find_value_columns(ws)

    idx_total = idx_by_label.get("00_総数")
    if idx_total is None:
        raise ValueError("census xlsxに '00_総数' 列(単位=人)が見つかりません。")

    idx_bands: Dict[int, int] = {}
    for n in range(1, 22):
        prefix = f"{n:02d}_"
        matches = [lbl for lbl in idx_by_label if lbl.startswith(prefix)]
        if len(matches) != 1:
            raise ValueError(
                f"census xlsxで年齢階級コード '{prefix}' に一致する列(単位=人)が"
                f"ちょうど1個ではありません(実測 {len(matches)} 個: {matches})。"
            )
        idx_bands[n] = idx_by_label[matches[0]]

    unknown_matches = [lbl for lbl in idx_by_label if lbl.startswith("22_")]
    if len(unknown_matches) != 1:
        raise ValueError(
            f"census xlsxで年齢「不詳」列(コード'22_')に一致する列(単位=人)が"
            f"ちょうど1個ではありません(実測 {len(unknown_matches)} 個)。"
        )
    idx_unknown = idx_by_label[unknown_matches[0]]

    plus65_matches = [
        lbl for lbl in idx_by_label if lbl.startswith("R3_") and "65" in lbl and "以上" in lbl
    ]
    if len(plus65_matches) != 1:
        raise ValueError(
            f"census xlsxで65歳以上(再掲)列に一致する列(単位=人)がちょうど1個では"
            f"ありません(実測 {len(plus65_matches)} 個: {plus65_matches})。"
        )
    idx_65plus = idx_by_label[plus65_matches[0]]

    return {
        "idx_total": idx_total,
        "idx_bands": idx_bands,
        "idx_unknown": idx_unknown,
        "idx_65plus": idx_65plus,
    }


class Entry:
    __slots__ = ("total", "bands", "unknown_direct", "pop65plus", "raw_name", "pref2020_raw")

    def __init__(self, total, bands, unknown_direct, pop65plus, raw_name, pref2020_raw):
        self.total = total
        self.bands = bands  # list[21]
        self.unknown_direct = unknown_direct
        self.pop65plus = pop65plus
        self.raw_name = raw_name
        self.pref2020_raw = pref2020_raw

    @property
    def unknown_derived(self) -> int:
        return self.total - sum(self.bands)


def load_census(xlsx_path: Path) -> Tuple[dict, dict, dict, List[str]]:
    """census xlsxを読み込み、市区町村・都道府県・全国の3レベルに分けて返す。

    戻り値: (muni_data, pref_data, national_data, errors)
      muni_data[muni_code][sex_label] -> Entry
      pref_data[pref_code][sex_label] -> Entry
      national_data[sex_label] -> Entry
    errors は想定外の地域識別コード等の致命的でない異常メモ(空なら問題なし)。
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb["b02_07"]
    cols = resolve_census_columns(ws)
    idx_total = cols["idx_total"]
    idx_bands = cols["idx_bands"]
    idx_unknown = cols["idx_unknown"]
    idx_65plus = cols["idx_65plus"]

    muni_data: Dict[str, Dict[str, Entry]] = {}
    pref_data: Dict[str, Dict[str, Entry]] = {}
    national_data: Dict[str, Entry] = {}
    errors: List[str] = []

    for row in ws.iter_rows(min_row=13, values_only=True):
        kokuseki = row[0]
        if kokuseki != "0_国籍総数":
            continue
        sex_raw = row[1]
        ident = row[2]
        pref2020_raw = row[6]
        code2020 = row[7]
        name_raw = row[8]

        sex = SEX_LABEL_MAP.get(sex_raw)
        if sex is None:
            errors.append(f"未知の男女区分: {sex_raw!r}(行の地域名={name_raw!r})")
            continue

        total = num(row[idx_total])
        bands = [num(row[idx_bands[n]]) for n in range(1, 22)]
        unknown_direct = num(row[idx_unknown])
        pop65plus = num(row[idx_65plus])
        entry = Entry(total, bands, unknown_direct, pop65plus, name_raw, pref2020_raw)

        if ident == "a":
            if code2020 == "00000":
                national_data[sex] = entry
            else:
                pref_code = str(code2020)[:2]
                pref_data.setdefault(pref_code, {})[sex] = entry
        elif ident in MUNI_IDENTS:
            muni_data.setdefault(str(code2020), {})[sex] = entry
        elif ident in ("1", "9"):
            # 1=政令市全体(区と重複)、9=2000年組み替え(現行市区町村と重複)。
            # 意図的に使わない(docstring参照)。
            continue
        else:
            errors.append(f"未知の地域識別コード: {ident!r}(2020年_地域コード={code2020!r})")

    return muni_data, pref_data, national_data, errors


def strip_name_prefix(raw_name: Optional[str]) -> str:
    """census の地域名(例 '0014_函館市')から先頭の連番プレフィックスを除く。"""
    if raw_name is None:
        return ""
    s = str(raw_name)
    if "_" in s:
        head, _, tail = s.partition("_")
        if head.isdigit():
            return tail
    return s


def strip_pref_prefix(raw_pref: Optional[str]) -> str:
    """census の '01_北海道' のような表記から都道府県名だけを取り出す。"""
    if raw_pref is None:
        return ""
    s = str(raw_pref)
    if "_" in s:
        _, _, tail = s.partition("_")
        return tail
    return s


# --- 市区町村→構想区域マッピングの構築 ---------------------------------------


def build_municipality_mapping(
    a38_geojson_path: Path,
    area_geo_join_path: Path,
    mie_csv_path: Path,
    muni_data: Dict[str, Dict[str, Entry]],
) -> List[dict]:
    with a38_geojson_path.open("r", encoding="utf-8") as f:
        geo = json.load(f)

    join_df = pd.read_csv(area_geo_join_path, dtype=str)
    matched = {
        row["geo_code"]: row for _, row in join_df[join_df["join_status"] == "matched"].iterrows()
    }

    # 都道府県名の全国共通ルックアップ(census自身の '01_北海道' 表記から作る)。
    pref_name_lookup: Dict[str, str] = {}
    for muni_code, sex_dict in muni_data.items():
        entry = sex_dict.get("total")
        if entry is None:
            continue
        pref_code = muni_code[:2]
        if pref_code not in pref_name_lookup:
            pref_name_lookup[pref_code] = strip_pref_prefix(entry.pref2020_raw)

    mapping_rows: List[dict] = []
    seen_muni_codes: Dict[str, str] = {}  # muni_code -> mapping_source(重複検出用)

    def add_row(muni_code: str, muni_name: str, area_code: str, area_name: str, source: str) -> None:
        if muni_code in seen_muni_codes:
            raise ValueError(
                f"市区町村コード {muni_code} が複数回マッピングされています"
                f"(既存: {seen_muni_codes[muni_code]!r}, 今回: {source!r})。"
            )
        seen_muni_codes[muni_code] = source
        pref_code = muni_code[:2]
        pref_name = pref_name_lookup.get(pref_code, "")
        mapping_rows.append(
            {
                "muni_code": muni_code,
                "muni_name": muni_name,
                "area_code": area_code,
                "area_name": area_name,
                "pref_code": pref_code,
                "pref_name": pref_name,
                "mapping_source": source,
            }
        )

    # --- A38b_001 経由(331件、三重県以外) ---------------------------------
    for feat in geo["features"]:
        props = feat["properties"]
        geo_code = props["A38b_003"]
        if geo_code not in matched:
            continue
        join_row = matched[geo_code]
        area_code = join_row["area_code"]
        area_name = join_row["area_name"]
        muni_codes_raw = props["A38b_001"]
        if not muni_codes_raw:
            continue
        for muni_code in str(muni_codes_raw).split(","):
            muni_code = muni_code.strip()
            if not muni_code:
                continue
            entry = muni_data.get(muni_code, {}).get("total")
            muni_name = strip_name_prefix(entry.raw_name) if entry is not None else ""
            add_row(muni_code, muni_name, area_code, area_name, "A38b_001")

    # --- 三重県: mie_area_municipalities.csv 経由 --------------------------
    mie_df = pd.read_csv(mie_csv_path, dtype=str)
    for _, row in mie_df.iterrows():
        add_row(
            row["muni_code"],
            row["muni_name"],
            row["area_code"],
            row["area_name"],
            "mie_area_municipalities",
        )

    return mapping_rows


# --- 集計 --------------------------------------------------------------------


def sum_entries(entries: List[Entry]) -> Tuple[int, List[int], int, int]:
    total = sum(e.total for e in entries)
    bands = [0] * 21
    for e in entries:
        for i, b in enumerate(e.bands):
            bands[i] += b
    unknown = sum(e.unknown_direct for e in entries)
    pop65plus = sum(e.pop65plus for e in entries)
    return total, bands, unknown, pop65plus


def age_row_values(total: int, bands: List[int], unknown: int, pop65plus: int) -> List[int]:
    return bands + [pop65plus, unknown, total]


# --- メイン ------------------------------------------------------------------


def parse_args(argv: List[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="年齢構成人口(issue #28)を整備する")
    parser.add_argument("--census-xlsx", type=Path, default=DEFAULT_CENSUS_XLSX)
    parser.add_argument(
        "--a38-geojson",
        type=Path,
        default=None,
        help=f"未指定なら $NEIGHBOR_REPO/{NEIGHBOR_A38_GEOJSON}",
    )
    parser.add_argument(
        "--area-geo-join",
        type=Path,
        default=None,
        help=f"未指定なら $NEIGHBOR_REPO/{NEIGHBOR_AREA_GEO_JOIN}",
    )
    parser.add_argument(
        "--mie-csv",
        type=Path,
        default=None,
        help=f"未指定なら $NEIGHBOR_REPO/{NEIGHBOR_MIE_CSV}",
    )
    parser.add_argument("--area-csv", type=Path, default=DEFAULT_AREA_CSV, help="既存 population_iryoken2.csv")
    parser.add_argument("--pref-csv", type=Path, default=DEFAULT_PREF_CSV, help="既存 population_prefecture.csv")
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    return parser.parse_args(argv)


def main(argv: List[str]) -> int:
    args = parse_args(argv)

    # 隣リポジトリ由来の3入力は、個別指定 → $NEIGHBOR_REPO の順に解決する。
    # どちらも無ければ入手手順を案内して止める(issue #51)。
    try:
        args.a38_geojson = lib_neighbor_repo.resolve(
            args.a38_geojson, NEIGHBOR_A38_GEOJSON, "--a38-geojson"
        )
        args.area_geo_join = lib_neighbor_repo.resolve(
            args.area_geo_join, NEIGHBOR_AREA_GEO_JOIN, "--area-geo-join"
        )
        args.mie_csv = lib_neighbor_repo.resolve(args.mie_csv, NEIGHBOR_MIE_CSV, "--mie-csv")
    except lib_neighbor_repo.NeighborRepoNotConfigured as exc:
        print(f"エラー: {exc}", file=sys.stderr)
        return 1

    # 隣リポジトリ由来の入力は、見つからないときに入手手順も併せて案内する。
    for label, path, hint in [
        ("census xlsx", args.census_xlsx, None),
        ("A38 geojson", args.a38_geojson, (NEIGHBOR_A38_GEOJSON, "--a38-geojson")),
        ("area_geo_join.csv", args.area_geo_join, (NEIGHBOR_AREA_GEO_JOIN, "--area-geo-join")),
        ("mie_area_municipalities.csv", args.mie_csv, (NEIGHBOR_MIE_CSV, "--mie-csv")),
        ("population_iryoken2.csv", args.area_csv, None),
        ("population_prefecture.csv", args.pref_csv, None),
    ]:
        if not path.exists():
            print(f"エラー: {label} が見つかりません: {path}", file=sys.stderr)
            if hint is not None:
                print(lib_neighbor_repo.guidance(*hint), file=sys.stderr)
            return 1

    print(f"census xlsx を読み込み中: {args.census_xlsx}")
    muni_data, pref_data, national_data, load_errors = load_census(args.census_xlsx)
    ok = True
    if load_errors:
        ok = False
        print(f"エラー: census読み込み中に想定外の値({len(load_errors)}件):")
        for e in load_errors[:50]:
            print(f"  {e}")

    print(f"  市区町村: {len(muni_data)}件 / 都道府県: {len(pref_data)}件 / 全国行: {len(national_data)}性別分")
    print()

    print("市区町村→構想区域マッピングを構築中...")
    mapping_rows = build_municipality_mapping(args.a38_geojson, args.area_geo_join, args.mie_csv, muni_data)
    mapping_rows.sort(key=lambda r: r["muni_code"])
    muni_code_to_area = {r["muni_code"]: r for r in mapping_rows}
    print(f"  マッピング行数: {len(mapping_rows)}件")
    print()

    audit: List[AuditRow] = []

    # --- 検算: 対応表の網羅性 -------------------------------------------------
    census_muni_codes = set(muni_data.keys())
    mapping_muni_codes = set(muni_code_to_area.keys())
    only_in_census = sorted(census_muni_codes - mapping_muni_codes)
    only_in_mapping = sorted(mapping_muni_codes - census_muni_codes)
    print("対応表の網羅性検算:")
    print(f"  census市区町村コード: {len(census_muni_codes)}件 / マッピング: {len(mapping_muni_codes)}件")
    if only_in_census:
        ok = False
        print(f"  エラー: censusにのみ存在({len(only_in_census)}件): {only_in_census[:30]}")
        for code in only_in_census:
            entry = muni_data.get(code, {}).get("total")
            name = strip_name_prefix(entry.raw_name) if entry else ""
            audit.append(AuditRow("muni_only_in_census", code, name, "", "", "", "mapping側に対応する行が無い"))
    if only_in_mapping:
        ok = False
        print(f"  エラー: マッピングにのみ存在({len(only_in_mapping)}件): {only_in_mapping[:30]}")
        for code in only_in_mapping:
            row = muni_code_to_area.get(code, {})
            audit.append(
                AuditRow("muni_only_in_mapping", code, row.get("muni_name", ""), "", "", "", "census側に対応する行が無い")
            )
    if not only_in_census and not only_in_mapping:
        print("  完全一致(片方にしか無いコードは無し)。")
    print()

    # --- 集計: 市区町村 -> 構想区域(339)/都道府県(合計チェック用) -----------------
    area_agg: Dict[str, Dict[str, Tuple[int, List[int], int, int]]] = {}  # area_code -> sex -> (total,bands,unknown,65plus)
    pref_from_areas: Dict[str, Dict[str, Tuple[int, List[int], int, int]]] = {}

    for sex in ("total", "male", "female"):
        by_area: Dict[str, List[Entry]] = {}
        for muni_code, row in muni_code_to_area.items():
            entry = muni_data.get(muni_code, {}).get(sex)
            if entry is None:
                continue
            by_area.setdefault(row["area_code"], []).append(entry)
        for area_code, entries in by_area.items():
            area_agg.setdefault(area_code, {})[sex] = sum_entries(entries)

    area_code_to_pref = {r["area_code"]: r["pref_code"] for r in mapping_rows}
    # area_codeごとのpref_codeは、マッピング行(muni単位)からarea_code->pref_codeを
    # 一意に取れるはず(1区域は1都道府県に属する)。念のため一意性を検算する。
    area_pref_sets: Dict[str, set] = {}
    for r in mapping_rows:
        area_pref_sets.setdefault(r["area_code"], set()).add(r["pref_code"])
    non_unique_areas = {a: p for a, p in area_pref_sets.items() if len(p) != 1}
    if non_unique_areas:
        ok = False
        print(f"エラー: 区域が単一の都道府県に属していません: {non_unique_areas}")

    for sex in ("total", "male", "female"):
        by_pref: Dict[str, List[Tuple[int, List[int], int, int]]] = {}
        for area_code, sex_dict in area_agg.items():
            if sex not in sex_dict:
                continue
            pref_code = area_code_to_pref.get(area_code)
            by_pref.setdefault(pref_code, []).append(sex_dict[sex])
        for pref_code, tuples in by_pref.items():
            totals = [t[0] for t in tuples]
            bandss = [t[1] for t in tuples]
            unknowns = [t[2] for t in tuples]
            plus65s = [t[3] for t in tuples]
            bands_sum = [0] * 21
            for b in bandss:
                for i, v in enumerate(b):
                    bands_sum[i] += v
            pref_from_areas.setdefault(pref_code, {})[sex] = (
                sum(totals),
                bands_sum,
                sum(unknowns),
                sum(plus65s),
            )

    # --- ハード検算: 全国検算(選択した市区町村行 == 全国行) --------------------
    print("全国検算(Σ選択市区町村行 == 全国行):")
    for sex in ("total", "male", "female"):
        entries = [sex_dict[sex] for sex_dict in muni_data.values() if sex in sex_dict]
        muni_total, muni_bands, muni_unknown, muni_65plus = sum_entries(entries)
        nat_entry = national_data.get(sex)
        if nat_entry is None:
            ok = False
            print(f"  エラー: 全国行に性別={sex}が見つかりません。")
            continue
        total_diff = muni_total - nat_entry.total
        band_diffs = [muni_bands[i] - nat_entry.bands[i] for i in range(21)]
        unknown_diff = muni_unknown - nat_entry.unknown_direct
        plus65_diff = muni_65plus - nat_entry.pop65plus
        bad_bands = [i + 1 for i, d in enumerate(band_diffs) if d != 0]
        if total_diff != 0 or bad_bands or unknown_diff != 0 or plus65_diff != 0:
            ok = False
            print(
                f"  エラー(sex={sex}): total_diff={total_diff}, 不一致バンド={bad_bands}, "
                f"unknown_diff={unknown_diff}, 65plus_diff={plus65_diff}"
            )
            audit.append(
                AuditRow(
                    "national_reconciliation",
                    "00000",
                    "全国",
                    nat_entry.total,
                    muni_total,
                    total_diff,
                    f"sex={sex}",
                )
            )
        else:
            print(f"  sex={sex}: 完全一致(total/21バンド/unknown/65plus)。")

    # Σ(339区域) == 全国行
    print("全国検算(Σ339区域 == 全国行):")
    for sex in ("total", "male", "female"):
        totals = [sex_dict[sex][0] for sex_dict in area_agg.values() if sex in sex_dict]
        n_areas_with_sex = len(totals)
        area_sum_total = sum(totals)
        nat_entry = national_data.get(sex)
        if nat_entry is None:
            ok = False
            print(f"  エラー(sex={sex}): 全国行が見つかりません。")
            continue
        diff = area_sum_total - nat_entry.total
        if n_areas_with_sex != 339 or diff != 0:
            ok = False
            print(
                f"  エラー(sex={sex}): 区域数={n_areas_with_sex}(期待339), "
                f"Σ区域合計={area_sum_total}, 全国行={nat_entry.total}, diff={diff}"
            )
        else:
            print(f"  sex={sex}: 339区域すべて揃い、合計が全国行と一致。")
    print()

    # --- ハード検算: 性別整合性(男+女==総数) ------------------------------------
    print("性別整合性検算(男+女==総数):")
    sex_bad = []
    for muni_code, sex_dict in muni_data.items():
        if "male" in sex_dict and "female" in sex_dict and "total" in sex_dict:
            diff = sex_dict["male"].total + sex_dict["female"].total - sex_dict["total"].total
            if diff != 0:
                sex_bad.append((muni_code, diff))
    for pref_code, sex_dict in pref_data.items():
        if "male" in sex_dict and "female" in sex_dict and "total" in sex_dict:
            diff = sex_dict["male"].total + sex_dict["female"].total - sex_dict["total"].total
            if diff != 0:
                sex_bad.append((f"pref:{pref_code}", diff))
    if "male" in national_data and "female" in national_data and "total" in national_data:
        diff = national_data["male"].total + national_data["female"].total - national_data["total"].total
        if diff != 0:
            sex_bad.append(("national", diff))
    if sex_bad:
        ok = False
        print(f"  エラー: 男+女 != 総数 が {len(sex_bad)}件: {sex_bad[:30]}")
        for code, diff in sex_bad:
            audit.append(AuditRow("sex_consistency", code, "", "", "", diff, "男+女 != 総数"))
    else:
        print("  すべての市区町村・都道府県・全国行で男+女==総数を確認。")
    print()

    # --- ハード検算: unknown_direct と derived(総数-Σバンド)の一致 -----------
    print("年齢不詳の整合性検算(総数-Σバンド == census '22_年齢不詳'列):")
    unknown_bad = []
    for muni_code, sex_dict in muni_data.items():
        for sex, entry in sex_dict.items():
            if entry.unknown_derived != entry.unknown_direct:
                unknown_bad.append((muni_code, sex, entry.unknown_derived, entry.unknown_direct))
            if entry.unknown_derived < 0:
                unknown_bad.append((muni_code, sex, entry.unknown_derived, "負値"))
    if unknown_bad:
        ok = False
        print(f"  エラー: {len(unknown_bad)}件: {unknown_bad[:30]}")
    else:
        print(f"  {sum(len(v) for v in muni_data.values())}件すべてで一致・非負を確認。")
    print()

    # --- ハード検算: pop65plus(census '(再掲)65歳以上'列) と Σ(65-69〜100歳以上バンド)の一致 --
    # census自身が持つ「(再掲)65歳以上」列は、5歳階級バンドの再集計ではなく
    # census側で独立に集計された値。今回の実測ではバンド合計と完全一致するが、
    # 将来 e-Stat が「(再掲)」の定義(例: 年齢不詳の按分方法)を変えた場合に
    # pop_65plus とバンド合計が食い違ったまま出力される事故を防ぐため、
    # 全market行(市区町村×性別・都道府県×性別・全国×性別)でハード検算する。
    print("65歳以上の整合性検算(census '(再掲)65歳以上' == Σ65-69〜100歳以上バンド):")
    plus65_bad = []
    for muni_code, sex_dict in muni_data.items():
        for sex, entry in sex_dict.items():
            band_sum_65plus = sum(entry.bands[13:])
            if entry.pop65plus != band_sum_65plus:
                plus65_bad.append((f"muni:{muni_code}", sex, entry.pop65plus, band_sum_65plus))
    for pref_code, sex_dict in pref_data.items():
        for sex, entry in sex_dict.items():
            band_sum_65plus = sum(entry.bands[13:])
            if entry.pop65plus != band_sum_65plus:
                plus65_bad.append((f"pref:{pref_code}", sex, entry.pop65plus, band_sum_65plus))
    for sex, entry in national_data.items():
        band_sum_65plus = sum(entry.bands[13:])
        if entry.pop65plus != band_sum_65plus:
            plus65_bad.append(("national", sex, entry.pop65plus, band_sum_65plus))
    n_65plus_checked = (
        sum(len(v) for v in muni_data.values())
        + sum(len(v) for v in pref_data.values())
        + len(national_data)
    )
    if plus65_bad:
        ok = False
        print(f"  エラー: {len(plus65_bad)}件(pop65plus != Σ65歳以上バンド): {plus65_bad[:30]}")
        for code, sex, expected, actual in plus65_bad:
            audit.append(
                AuditRow("pop65plus_vs_band_sum", code, "", expected, actual, expected - actual, f"sex={sex}")
            )
    else:
        print(f"  {n_65plus_checked}件すべてで一致を確認(市区町村×性別 + 都道府県×性別 + 全国×性別)。")
    print()

    # --- 都道府県ロールアップ検算(Σ339区域(都道府県別) == censusの都道府県行) --
    print("都道府県ロールアップ検算(Σ区域(都道府県別) == censusの都道府県行):")
    pref_rollup_bad = []
    for pref_code, sex_dict in pref_from_areas.items():
        census_sex_dict = pref_data.get(pref_code, {})
        for sex, (total, bands, unknown, plus65) in sex_dict.items():
            census_entry = census_sex_dict.get(sex)
            if census_entry is None:
                pref_rollup_bad.append((pref_code, sex, "census側に都道府県行なし"))
                continue
            diff = total - census_entry.total
            if diff != 0:
                pref_rollup_bad.append((pref_code, sex, diff))
                audit.append(
                    AuditRow(
                        "pref_rollup",
                        pref_code,
                        "",
                        census_entry.total,
                        total,
                        diff,
                        f"sex={sex}",
                    )
                )
    if pref_rollup_bad:
        ok = False
        print(f"  エラー: {len(pref_rollup_bad)}件: {pref_rollup_bad[:30]}")
    else:
        print("  すべての都道府県・性別でΣ区域合計とcensus都道府県行が一致。")
    print()

    if not ok:
        print("検算NG: 上記のいずれかのハード検算に失敗しました。監査CSVを書き出して終了します。")
        write_csv(
            args.out_dir / "population_age_audit.csv",
            ["check", "code", "name", "expected", "actual", "diff", "note"],
            [a.as_row() for a in audit],
        )
        return 1

    # --- 出力: municipality_to_iryoken2.csv --------------------------------
    mapping_out_path = args.out_dir / "municipality_to_iryoken2.csv"
    write_csv(
        mapping_out_path,
        ["muni_code", "muni_name", "area_code", "area_name", "pref_code", "pref_name", "mapping_source"],
        [
            [r["muni_code"], r["muni_name"], r["area_code"], r["area_name"], r["pref_code"], r["pref_name"], r["mapping_source"]]
            for r in mapping_rows
        ],
    )
    print(f"生成: {mapping_out_path}({len(mapping_rows)}件)")

    # --- 出力: population_iryoken2.csv(既存列+年齢列、性別計) ------------------
    existing_area = pd.read_csv(args.area_csv, dtype=str, usecols=BASE_AREA_COLUMNS)
    existing_area = existing_area[BASE_AREA_COLUMNS]
    area_rows_out = []
    area_name_by_code = {r["area_code"]: r["area_name"] for r in mapping_rows}
    census_vs_existing_area: List[Tuple[str, int, int]] = []
    for _, row in existing_area.iterrows():
        area_code = row["area_code"]
        sex_dict = area_agg.get(area_code, {})
        if "total" not in sex_dict:
            print(f"エラー: 区域 {area_code} に年齢データがありません。", file=sys.stderr)
            return 1
        total, bands, unknown, plus65 = sex_dict["total"]
        age_values = age_row_values(total, bands, unknown, plus65)
        area_rows_out.append(list(row[BASE_AREA_COLUMNS]) + age_values)
        census_vs_existing_area.append((area_code, int(row["population_2020"]), total))
    write_csv(args.out_dir / "population_iryoken2.csv", BASE_AREA_COLUMNS + AGE_ALL_COLUMNS, area_rows_out)
    print(f"更新: {args.out_dir / 'population_iryoken2.csv'}({len(area_rows_out)}件、年齢列{len(AGE_ALL_COLUMNS)}列追加)")

    # --- 出力: population_prefecture.csv(既存列+年齢列、性別計、census自身の都道府県行) --
    existing_pref = pd.read_csv(args.pref_csv, dtype=str, usecols=BASE_PREF_COLUMNS)
    existing_pref = existing_pref[BASE_PREF_COLUMNS]
    pref_rows_out = []
    census_vs_existing_pref: List[Tuple[str, int, int]] = []
    for _, row in existing_pref.iterrows():
        pref_code = row["pref_code"]
        sex_dict = pref_data.get(pref_code, {})
        if "total" not in sex_dict:
            print(f"エラー: 都道府県 {pref_code} に年齢データがありません。", file=sys.stderr)
            return 1
        entry = sex_dict["total"]
        age_values = age_row_values(entry.total, entry.bands, entry.unknown_direct, entry.pop65plus)
        pref_rows_out.append(list(row[BASE_PREF_COLUMNS]) + age_values)
        census_vs_existing_pref.append((pref_code, int(row["population_2020"]), entry.total))
    write_csv(args.out_dir / "population_prefecture.csv", BASE_PREF_COLUMNS + AGE_ALL_COLUMNS, pref_rows_out)
    print(f"更新: {args.out_dir / 'population_prefecture.csv'}({len(pref_rows_out)}件、年齢列{len(AGE_ALL_COLUMNS)}列追加)")

    # --- 出力: population_iryoken2_age_sex.csv -------------------------------
    area_sex_rows = []
    for _, row in existing_area.iterrows():
        area_code = row["area_code"]
        for sex_label, sex_out in (("male", "male"), ("female", "female")):
            total, bands, unknown, plus65 = area_agg[area_code][sex_label]
            age_values = age_row_values(total, bands, unknown, plus65)
            area_sex_rows.append(
                [area_code, row["area_name"], row["pref_code"], row["pref_name"], sex_out] + age_values
            )
    write_csv(
        args.out_dir / "population_iryoken2_age_sex.csv",
        ["area_code", "area_name", "pref_code", "pref_name", "sex"] + AGE_ALL_COLUMNS,
        area_sex_rows,
    )
    print(f"生成: {args.out_dir / 'population_iryoken2_age_sex.csv'}({len(area_sex_rows)}件)")

    # --- 出力: population_prefecture_age_sex.csv -----------------------------
    pref_sex_rows = []
    for _, row in existing_pref.iterrows():
        pref_code = row["pref_code"]
        for sex_label, sex_out in (("male", "male"), ("female", "female")):
            entry = pref_data[pref_code][sex_label]
            age_values = age_row_values(entry.total, entry.bands, entry.unknown_direct, entry.pop65plus)
            pref_sex_rows.append([pref_code, row["pref_name"], sex_out] + age_values)
    write_csv(
        args.out_dir / "population_prefecture_age_sex.csv",
        ["pref_code", "pref_name", "sex"] + AGE_ALL_COLUMNS,
        pref_sex_rows,
    )
    print(f"生成: {args.out_dir / 'population_prefecture_age_sex.csv'}({len(pref_sex_rows)}件)")
    print()

    # --- 監査: pop_total_census vs 既存 population_2020(0.5%超で失敗) --------
    print("pop_total_census と既存 population_2020 の差の監査:")
    max_rel_diff = 0.0
    hard_fail_diffs: List[Tuple[str, str, int, int]] = []
    for level, records in (("area", census_vs_existing_area), ("pref", census_vs_existing_pref)):
        n_diff = 0
        level_max_abs_diff = 0
        for code, existing_pop, census_total in records:
            diff = census_total - existing_pop
            if diff != 0:
                n_diff += 1
                level_max_abs_diff = max(level_max_abs_diff, abs(diff))
                rel_diff = abs(diff) / existing_pop if existing_pop else float("inf")
                max_rel_diff = max(max_rel_diff, rel_diff)
                name = area_name_by_code.get(code, "") if level == "area" else ""
                audit.append(
                    AuditRow(
                        f"{level}_total_vs_population_2020",
                        code,
                        name,
                        existing_pop,
                        census_total,
                        diff,
                        f"相対差={rel_diff:.4%}" if existing_pop else "existing pop=0",
                    )
                )
                if existing_pop and rel_diff > 0.005:
                    hard_fail_diffs.append((level, code, existing_pop, census_total))
        print(f"  {level}: 差がある件数={n_diff} / 最大絶対差={level_max_abs_diff}")
    print(f"  最大相対差: {max_rel_diff:.4%}")
    if hard_fail_diffs:
        ok = False
        print(f"  エラー: 相対差0.5%超が{len(hard_fail_diffs)}件: {hard_fail_diffs[:20]}")
    print()

    # --- 監査CSVは常に書く(空でもヘッダのみ) ----------------------------------
    write_csv(
        args.out_dir / "population_age_audit.csv",
        ["check", "code", "name", "expected", "actual", "diff", "note"],
        [a.as_row() for a in audit],
    )
    print(f"生成: {args.out_dir / 'population_age_audit.csv'}({len(audit)}行)")
    print()

    if not ok:
        print("検算NG: pop_total_census vs population_2020 の相対差0.5%超があります。")
        return 1

    print("検算OK: すべてのハード検算に合格しました。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
